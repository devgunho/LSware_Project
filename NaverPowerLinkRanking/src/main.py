import sys
import os
import json
import re
import chardet
import time
import threading
from time import sleep

import ErrorPopup_Tkinter as ePopup
import FileOpen_easygui as fopen
import URL_KeywordParser as urlParser
import MakeCSV

from tkinter import *
from tkinter.ttk import *
import tkinter.messagebox

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtGui, QtWidgets

from urllib.parse import quote_plus
from openpyxl import load_workbook as load_workbook

url_col = []
keyword_col = []
ranking_col = []
download1Cnt = 0
searchBtnCnt = 0
PageCnt = 1

lineCnt = 0     # CSV Data Line Counter
multiPageCnt = 1
url_col2 = []
keyword_col2 = []
ranking_col2 = []
download2Cnt = 0
totalLines = 0
multiSearchFilePath = ''


class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("PowerLink Ranking Program")
        Dialog.resize(692, 892)
        self.tabWidget = QtWidgets.QTabWidget(Dialog)
        self.tabWidget.setGeometry(QtCore.QRect(10, 10, 671, 861))
        self.tabWidget.setObjectName("tabWidget")

        # Basic Tab
        self.Basic_Tab = QtWidgets.QWidget()
        self.Basic_Tab.setObjectName("Basic_Tab")
        self.label_Keyword = QtWidgets.QLabel(self.Basic_Tab)
        self.label_Keyword.setGeometry(QtCore.QRect(50, 70, 70, 20))
        self.label_Keyword.setObjectName("label_Keyword")
        self.label_URL = QtWidgets.QLabel(self.Basic_Tab)
        self.label_URL.setGeometry(QtCore.QRect(60, 30, 71, 21))
        self.label_URL.setObjectName("label_URL")

        self.Input_URL = QtWidgets.QLineEdit(self.Basic_Tab)
        self.Input_URL.setGeometry(QtCore.QRect(150, 20, 491, 31))
        self.Input_URL.setObjectName("Input_URL")
        self.Input_URL.setText("검색하려는 URL을 입력해주세요.")

        self.Input_Keyword = QtWidgets.QLineEdit(self.Basic_Tab)
        self.Input_Keyword.setGeometry(QtCore.QRect(150, 70, 491, 31))
        self.Input_Keyword.setObjectName("Input_Keyword")
        self.Input_Keyword.setText("해당 URL이 조회되기 위한 키워드를 입력해주세요.")

        self.Search = QtWidgets.QPushButton(self.Basic_Tab)
        self.Search.setGeometry(QtCore.QRect(520, 120, 121, 23))
        self.Search.setObjectName("Search")
        self.Search.clicked.connect(self.SearchBtnClicked)

        self.ResultTable = QtWidgets.QTableWidget(self.Basic_Tab)
        self.ResultTable.setGeometry(QtCore.QRect(20, 190, 621, 581))
        self.ResultTable.setObjectName("ResultTable")
        self.ResultTable.setColumnCount(3)
        self.ResultTable.setRowCount(25)
        column_headers = ['URL 주소', '키워드', '순위']
        self.ResultTable.setHorizontalHeaderLabels(column_headers)
        self.ResultTable.resizeColumnsToContents()
        self.ResultTable.resizeRowsToContents()

        self.Download = QtWidgets.QPushButton(self.Basic_Tab)
        self.Download.setGeometry(QtCore.QRect(520, 150, 121, 23))
        self.Download.setObjectName("Download")
        self.Download.clicked.connect(self.DownloadBtnClicked)

        self.Next = QtWidgets.QPushButton(self.Basic_Tab)
        self.Next.setGeometry(QtCore.QRect(420, 790, 20, 20))
        self.Next.setObjectName("Next")
        self.Next.clicked.connect(self.NextBtnClicked)

        self.pageLabel = QtWidgets.QLabel(self.Basic_Tab)
        self.pageLabel.setGeometry(QtCore.QRect(295, 790, 75, 23))
        self.pageLabel.setObjectName("pageLabel1")
        self.Previous = QtWidgets.QPushButton(self.Basic_Tab)
        self.Previous.setGeometry(QtCore.QRect(200, 790, 20, 20))
        self.Previous.setObjectName("Previous")
        self.Previous.clicked.connect(self.PreviousBtnClicked)
        self.tabWidget.addTab(self.Basic_Tab, "")

        self.resetTable = QtWidgets.QPushButton(self.Basic_Tab)
        self.resetTable.setGeometry(QtCore.QRect(390, 150, 121, 23))
        self.resetTable.setObjectName("Reset")
        self.resetTable.clicked.connect(self.ResetBtnClicked)

        # Multi_Tab
        self.Multi_Tab = QtWidgets.QWidget()
        self.Multi_Tab.setObjectName("Multi_Tab")
        self.UploadLabel = QtWidgets.QLabel(self.Multi_Tab)
        self.UploadLabel.setGeometry(QtCore.QRect(40, 50, 75, 23))
        self.UploadLabel.setObjectName("Upload")

        self.LocalPath = QtWidgets.QLineEdit(self.Multi_Tab)
        self.LocalPath.setGeometry(QtCore.QRect(130, 40, 501, 31))
        self.LocalPath.setObjectName("LocalPath")
        self.LocalPath.setText(".xlsx 파일의 절대주소를 입력해주세요.")

        self.FileOpen = QtWidgets.QPushButton(self.Multi_Tab)
        self.FileOpen.setGeometry(QtCore.QRect(520, 90, 121, 23))
        self.FileOpen.setObjectName("FileOpen")
        self.FileOpen.clicked.connect(self.FileOpenBtnClicked)

        self.Search2 = QtWidgets.QPushButton(self.Multi_Tab)
        self.Search2.setGeometry(QtCore.QRect(520, 120, 121, 23))
        self.Search2.setObjectName("Search2")
        self.Search2.clicked.connect(self.Search2BtnClicked)

        self.Download2 = QtWidgets.QPushButton(self.Multi_Tab)
        self.Download2.setGeometry(QtCore.QRect(520, 150, 121, 23))
        self.Download2.setObjectName("Download2")
        self.Download2.clicked.connect(self.Download2BtnClicked)

        self.Previous2 = QtWidgets.QPushButton(self.Multi_Tab)
        self.Previous2.setGeometry(QtCore.QRect(200, 790, 20, 20))
        self.Previous2.setObjectName("Previous2")
        self.Previous2.clicked.connect(self.Previous2BtnClicked)

        self.pageLabel2 = QtWidgets.QLabel(self.Multi_Tab)
        self.pageLabel2.setGeometry(QtCore.QRect(295, 790, 75, 23))
        self.pageLabel2.setObjectName("pageLabel2")

        self.Next22 = QtWidgets.QPushButton(self.Multi_Tab)
        self.Next22.setGeometry(QtCore.QRect(420, 790, 20, 20))
        self.Next22.setObjectName("Next22")
        self.Next22.clicked.connect(self.Next2BtnClicked)

        self.ResultTable2 = QtWidgets.QTableWidget(self.Multi_Tab)
        self.ResultTable2.setGeometry(QtCore.QRect(20, 190, 621, 581))
        self.ResultTable2.setObjectName("ResultTable2")
        self.ResultTable2.setColumnCount(3)
        self.ResultTable2.setRowCount(25)
        self.ResultTable2.setHorizontalHeaderLabels(column_headers)
        self.ResultTable2.resizeColumnsToContents()
        self.ResultTable2.resizeRowsToContents()

        self.resetTable2 = QtWidgets.QPushButton(self.Multi_Tab)
        self.resetTable2.setGeometry(QtCore.QRect(390, 150, 121, 23))
        self.resetTable2.setObjectName("Reset2")
        self.resetTable2.clicked.connect(self.ResetBtn2Clicked)

        self.tabWidget.addTab(self.Multi_Tab, "")

        self.retranslateUi(Dialog)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate(
            "Dialog", "네이버 파워링크 순위 검색"))
        self.label_Keyword.setText(_translate("Dialog", "검색 키워드"))
        self.label_URL.setText(_translate("Dialog", "URL 주소"))
        self.Search.setText(_translate("Dialog", "검색"))
        self.Download.setText(_translate("Dialog", "다운로드 (.csv)"))
        self.Next.setText(_translate("Dialog", ">"))
        self.pageLabel.setText(_translate("Dialog", "1"))
        self.Previous.setText(_translate("Dialog", "<"))
        self.resetTable.setText(_translate("Dialog", "초기화"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(
            self.Basic_Tab), _translate("Dialog", "기본"))
        self.UploadLabel.setText(_translate("Dialog", "파일 경로"))
        self.Search2.setText(_translate("Dialog", "검색"))
        self.Download2.setText(_translate("Dialog", "다운로드 (.csv)"))
        self.Next22.setText(_translate("Dialog", ">"))
        self.pageLabel2.setText(_translate("Dialog", "1"))
        self.Previous2.setText(_translate("Dialog", "<"))
        self.resetTable2.setText(_translate("Dialog", "초기화"))
        self.FileOpen.setText(_translate("Dialog", "열기"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(
            self.Multi_Tab), _translate("Dialog", "대량조회"))

    def SearchBtnClicked(self):
        global searchBtnCnt, PageCnt
        searchBtnCnt += 1
        #print(searchBtnCnt, PageCnt)    # Debuging Point

        PageCntStr = ''
        if (searchBtnCnt % 25 == 0):
            for i in range(1, searchBtnCnt//25+1):
                PageCntStr += str(i)
                PageCntStr += str(' ')
        else:
            for i in range(1, searchBtnCnt//25+2):
                PageCntStr += str(i)
                PageCntStr += str(' ')
        self.pageLabel.setText(PageCntStr)

        keyword = self.Input_Keyword.text()
        searchUrl = self.Input_URL.text()
        urlResult, keywordResult, rankingResult = urlParser.PowerLinkPaser(
            searchUrl, keyword)
        url_col.extend(urlResult)
        keyword_col.extend(keywordResult)
        ranking_col.extend(rankingResult)

        tableTempData = {
            'url_col': url_col,
            'keyword_col': keyword_col,
            'ranking_col': ranking_col
        }
        column_idx_lookup = {'url_col': 0, 'keyword_col': 1, 'ranking_col': 2}

        self.ResultTable.setRowCount(0)
        self.ResultTable.setRowCount(25)
        for k, v in tableTempData.items():
            col = column_idx_lookup[k]
            for row, val in enumerate(v):
                if(row >= (PageCnt-1)*25):
                    item = QTableWidgetItem(val)
                    self.ResultTable.setItem(row-(PageCnt-1)*25, col, item)
        self.ResultTable.resizeColumnsToContents()
        self.ResultTable.resizeRowsToContents()

    def FileOpenBtnClicked(self):
        absPath = fopen.OpenWinFileExplorer()
        fileExtension = os.path.splitext(absPath)[1]
        #print(fileExtension)
        if (fileExtension == '.csv'):
            print("Load Ok!")
            pass
        elif (fileExtension == '.xlsx'):
            print("Load Ok!")
            pass
        else:
            ePopup.FileLoadError()
        self.LocalPath.setText(absPath)

    def Search2BtnClicked(self):
        keyword_col2.clear()
        ranking_col2.clear()
        url_col2.clear()
        self.ResultTable2.setRowCount(0)
        self.ResultTable2.setRowCount(25)
        self.ResultTable2.resizeColumnsToContents()
        self.ResultTable2.resizeRowsToContents()
        global multiSearchFilePath
        multiSearchFilePath = self.LocalPath.text()
        global lineCnt
        if (os.path.splitext(multiSearchFilePath)[1] == '.csv'):
            try:
                f = open(multiSearchFilePath, 'r', encoding='utf-8')
            except OSError:
                #print('cannot open : ', multiSearchFilePath)
                ePopup.loadWrongPath(multiSearchFilePath)
                pass
            else:
                lineCnt = 0
                reading = csv.reader(f)
                for line in reading:
                    try:
                        if (line[0] == 'URL') & (line[1] == 'KEYWORD'):
                            continue
                    except:
                        ePopup.loadWrongForm(multiSearchFilePath)
                        break
                    else:
                        lineCnt += 1
                        searchUrl = line[0]
                        keyword = line[1]
                        urlResult, keywordResult, rankingResult = urlParser.PowerLinkPaser(
                            searchUrl, keyword)
                        url_col2.extend(urlResult)
                        keyword_col2.extend(keywordResult)
                        ranking_col2.extend(rankingResult)
                #print(lineCnt)

                PageCntStr = '1 '
                if((lineCnt//25) > 0):
                    for i in range(2, lineCnt//25+2):
                        PageCntStr += str(i)
                        PageCntStr += str(' ')
                #print(PageCntStr)
                self.pageLabel2.setText(PageCntStr)

                tableTempData = {
                    'url_col': url_col2,
                    'keyword_col': keyword_col2,
                    'ranking_col': ranking_col2
                }
                column_idx_lookup = {'url_col': 0,
                                     'keyword_col': 1, 'ranking_col': 2}

                for k, v in tableTempData.items():
                    col = column_idx_lookup[k]
                    for row, val in enumerate(v):
                        item = QTableWidgetItem(val)
                        print(row,col,item)
                        self.ResultTable2.setItem(row, col, item)

                self.ResultTable2.resizeColumnsToContents()
                self.ResultTable2.resizeRowsToContents()

                f.close()
        else:
            ProgressApp = tkApp()
            ProgressApp.mainloop()

            PageCntStr = '1 '
            if((lineCnt//25) > 0):
                for i in range(2, lineCnt//25+2):
                    PageCntStr += str(i)
                    PageCntStr += str(' ')
            #print(PageCntStr)
            self.pageLabel2.setText(PageCntStr)

            tableTempData = {
                'url_col': url_col2,
                'keyword_col': keyword_col2,
                'ranking_col': ranking_col2
            }
            column_idx_lookup = {'url_col': 0,
                                 'keyword_col': 1, 'ranking_col': 2}

            for k, v in tableTempData.items():
                col = column_idx_lookup[k]
                for row, val in enumerate(v):
                    item = QTableWidgetItem(val)
                    self.ResultTable2.setItem(row, col, item)
            self.ResultTable2.resizeColumnsToContents()
            self.ResultTable2.resizeRowsToContents()

    def DownloadBtnClicked(self):
        tableTempData = {
            'URL': url_col,
            'KEYWORD': keyword_col,
            'RANKING': ranking_col
        }
        #print(tableTempData)

        current_path = os.path.dirname(os.path.realpath(__file__))
        download_path = current_path + r'\SearchResults'
        if not os.path.exists(download_path):
            os.makedirs(download_path)

        global download1Cnt, searchBtnCnt, PageCnt
        if(searchBtnCnt == 0):
            ePopup.NoneTableData()
        else:
            download1Cnt += 1
            output_file_name = download_path + '\SearchResults' + \
                str(download1Cnt) + quote_plus(".csv")
            MakeCSV.MakeDownloadCSV(output_file_name, tableTempData)
            url_col.clear()
            keyword_col.clear()
            ranking_col.clear()
            self.ResultTable.setRowCount(0)
            searchBtnCnt = 0
            PageCnt = 1
            PageCntStr = ''
            for i in range(1, searchBtnCnt//25+2):
                PageCntStr += str(i)
                PageCntStr += str(' ')
            self.pageLabel.setText(PageCntStr)
            self.ResultTable.setRowCount(25)
            self.ResultTable.resizeColumnsToContents()
            self.ResultTable.resizeRowsToContents()

    def Download2BtnClicked(self):
        tableTempData = {
            'URL': url_col2,
            'KEYWORD': keyword_col2,
            'RANKING': ranking_col2
        }
        #print(tableTempData)

        current_path = os.path.dirname(os.path.realpath(__file__))
        download_path = current_path + r'\MultiSearchResults'
        if not os.path.exists(download_path):
            os.makedirs(download_path)

        global download2Cnt, lineCnt
        if(lineCnt == 0):
            ePopup.NoneTableData()
        else:
            download2Cnt += 1
            output_file_name = download_path + '\MultiSearchResults' + \
                str(download2Cnt) + quote_plus(".csv")
            MakeCSV.MakeDownloadCSV(output_file_name, tableTempData)
        url_col2.clear()
        keyword_col2.clear()
        ranking_col2.clear()
        self.ResultTable2.setRowCount(0)
        self.ResultTable2.setRowCount(25)
        self.ResultTable2.resizeColumnsToContents()
        self.ResultTable2.resizeRowsToContents()

    def NextBtnClicked(self):
        global PageCnt
        if(searchBtnCnt > (25*PageCnt)):
            self.ResultTable.setRowCount(0)
            self.ResultTable.setRowCount(25)
            tableTempData = {
                'url_col': url_col,
                'keyword_col': keyword_col,
                'ranking_col': ranking_col
            }
            column_idx_lookup = {'url_col': 0,
                                 'keyword_col': 1, 'ranking_col': 2}

            for k, v in tableTempData.items():
                col = column_idx_lookup[k]
                for row, val in enumerate(v):
                    if(row >= (25*PageCnt)):
                        item = QTableWidgetItem(val)
                        self.ResultTable.setItem(row-PageCnt*25, col, item)
            self.ResultTable.resizeColumnsToContents()
            self.ResultTable.resizeRowsToContents()
            PageCnt += 1
        else:
            ePopup.PageBtnError()

    def Next2BtnClicked(self):
        #print("Next2BtnClicked")
        global multiPageCnt
        if(lineCnt > (25*multiPageCnt)):
            self.ResultTable2.setRowCount(0)
            self.ResultTable2.setRowCount(25)
            tableTempData = {
                'url_col': url_col2,
                'keyword_col': keyword_col2,
                'ranking_col': ranking_col2
            }
            column_idx_lookup = {'url_col': 0,
                                 'keyword_col': 1, 'ranking_col': 2}

            for k, v in tableTempData.items():
                col = column_idx_lookup[k]
                for row, val in enumerate(v):
                    if(row >= (25*multiPageCnt)):
                        item = QTableWidgetItem(val)
                        self.ResultTable2.setItem(
                            row-multiPageCnt*25, col, item)
            self.ResultTable2.resizeColumnsToContents()
            self.ResultTable2.resizeRowsToContents()
            multiPageCnt += 1
        else:
            ePopup.PageBtnError()

    def PreviousBtnClicked(self):
        global PageCnt
        if(PageCnt != 1):
            PageCnt -= 1
            self.ResultTable.setRowCount(0)
            self.ResultTable.setRowCount(25)
            tableTempData = {
                'url_col': url_col,
                'keyword_col': keyword_col,
                'ranking_col': ranking_col
            }
            column_idx_lookup = {'url_col': 0,
                                 'keyword_col': 1, 'ranking_col': 2}

            for k, v in tableTempData.items():
                col = column_idx_lookup[k]
                for row, val in enumerate(v):
                    if(row >= (PageCnt-1)*25):
                        item = QTableWidgetItem(val)
                        self.ResultTable.setItem(row-(PageCnt-1)*25, col, item)
            self.ResultTable.resizeColumnsToContents()
            self.ResultTable.resizeRowsToContents()
        else:
            ePopup.PageBtnError()

    def Previous2BtnClicked(self):
        #print("Previous2BtnClicked")
        global multiPageCnt
        if(multiPageCnt != 1):
            multiPageCnt -= 1
            self.ResultTable2.setRowCount(0)
            self.ResultTable2.setRowCount(25)
            tableTempData = {
                'url_col': url_col2,
                'keyword_col': keyword_col2,
                'ranking_col': ranking_col2
            }
            column_idx_lookup = {'url_col': 0,
                                 'keyword_col': 1, 'ranking_col': 2}

            for k, v in tableTempData.items():
                col = column_idx_lookup[k]
                for row, val in enumerate(v):
                    if(row >= (multiPageCnt-1)*25):
                        item = QTableWidgetItem(val)
                        self.ResultTable2.setItem(
                            row-(multiPageCnt-1)*25, col, item)
            self.ResultTable2.resizeColumnsToContents()
            self.ResultTable2.resizeRowsToContents()
        else:
            ePopup.PageBtnError()

    def ResetBtnClicked(self):
        self.ResultTable.setRowCount(0)
        self.ResultTable.setRowCount(25)
        self.ResultTable.resizeColumnsToContents()
        self.ResultTable.resizeRowsToContents()
        global download1Cnt, searchBtnCnt, PageCnt
        global url_col, keyword_col, ranking_col
        url_col.clear()
        keyword_col.clear()
        ranking_col.clear()
        download1Cnt = 0
        searchBtnCnt = 0
        PageCnt = 1

    def ResetBtn2Clicked(self):
        self.ResultTable2.setRowCount(0)
        self.ResultTable2.setRowCount(25)
        self.ResultTable2.resizeColumnsToContents()
        self.ResultTable2.resizeRowsToContents()
        global download2Cnt, multiPageCnt, lineCnt
        global url_col2, keyword_col2, ranking_col2
        url_col2.clear()
        keyword_col2.clear()
        ranking_col2.clear()
        download2Cnt = 0
        lineCnt = 0
        multiPageCnt = 1


# Progress Check Thread
progress_var = 0


class tkApp(Tk):
    global totalLines

    def __init__(self):
        super().__init__()
        self.title("진행상황")
        self.geometry('{}x{}'.format(400, 150))
        self.cautionText = tkinter.StringVar()
        self.cautionText.set("\n검색 진행 중 종료하지 마세요.\n")
        self.cautionLabel = Label(self, textvariable=self.cautionText)
        self.cautionLabel.pack()
        self.pgText = tkinter.StringVar()
        self.pgText.set("\n진행 : ( 0/0 )")
        self.theLabel = Label(self, textvariable=self.pgText)
        self.theLabel.pack()
        self.timeText = tkinter.StringVar()
        self.timeText.set("진행시간 : ")
        self.timeLabel = Label(self, textvariable=self.timeText)
        self.timeLabel.pack()
        self.checkBtn = Button(
            self, text='확인', state='disabled', takefocus=False, command=self.destroy)
        self.checkBtn.pack()

        self.progress = tkinter.ttk.Progressbar(
            self, variable=progress_var, mode="determinate")
        self.progress.pack(fill=X, expand=1)

        self.PgChanger()

    def PgChanger(self):
        wrongFormError = 0
        startTime = time.time()
        global lineCnt, multiSearchFilePath
        try:
            load_wb = load_workbook(multiSearchFilePath, data_only=True)
        except:
            #print('cannot open : ', multiSearchFilePath)
            ePopup.loadWrongPath(multiSearchFilePath)
        else:
            sheet = load_wb.worksheets[0]
            lineCnt = 0
            xlData = []
            for row in sheet.rows:
                xlData.append([row[0].value, row[1].value])
            for i, dt in enumerate(xlData):
                if(i == 0) & ((dt[0] != 'URL') | (dt[1] != 'KEYWORD')):
                    #print("Wrong Form!")
                    self.checkBtn['takefocus'] = TRUE
                    self.checkBtn['state'] = 'active'
                    self.update()
                    wrongFormError = 1
                    ePopup.loadWrongForm(multiSearchFilePath)
                    xlData.clear()
                    break
                else:
                    del xlData[0]
                    break
            totalLines = len(xlData)
            if(totalLines != 0):
                self.progress['maximum'] = totalLines-1

            global progress_var
            for i, dt in enumerate(xlData):
                lineCnt = i
                #print(i+1, dt[0], dt[1])    # check
                urlResult, keywordResult, rankingResult = urlParser.PowerLinkPaser(
                    dt[0], dt[1])
                url_col2.extend(urlResult)
                keyword_col2.extend(keywordResult)
                ranking_col2.extend(rankingResult)

                progress_var = lineCnt
                self.progress['value'] = progress_var
                progressText = '\n진행 : ('+str(lineCnt+1) + \
                    '/'+str(totalLines)+')'
                endTime = time.time() - startTime
                endTime = round(endTime, 2)
                tempTimeText = '진행시간 : ' + str(endTime)+' 초'
                self.timeText.set(tempTimeText)
                self.pgText.set(progressText)
                self.progress.update()
            load_wb.close()
            if(wrongFormError != 1):
                self.checkBtn['takefocus'] = TRUE
                self.checkBtn['state'] = 'active'


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()

    sys.exit(app.exec_())

t = threading.Thread(target=main)
t.start()